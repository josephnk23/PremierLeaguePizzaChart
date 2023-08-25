from urllib.request import urlopen
from bs4 import BeautifulSoup
import re
from lxml import html
import pandas as pd
import openpyxl
import unicodedata
from mplsoccer import PyPizza, FontManager, add_image
from PIL import Image, ImageDraw, ImageOps
import matplotlib.pyplot as plt
import time
import requests
import sqlite3
import io
import os


def getReports():
  #open the url and extract the html and write it to a new file
    html = urlopen('https://fbref.com/en/comps/9/stats/Premier-League-Stats')
    bs = BeautifulSoup(html, 'html.parser').encode().decode()

    with open("data1.html", "w") as file:
        file.write(str(bs))
        
#open the html file and replace the comment part with 
 #void so that all the tables can now be accessed
    with open("data1.html",'r') as html1:
        bs2 = BeautifulSoup(html1, 'html.parser')
        text_of_bs2 = str(bs2)
        usable_bs2 = text_of_bs2.replace("<!--"," ").replace("-->", " ")
        with open("data2.html", 'w') as commentless_content:
            commentless_content.write(usable_bs2)

#create an excel sheet and add links to player profiles so as if fetch it if need be
    with open('data2.html', 'r') as main_page:
      bs3 = BeautifulSoup(main_page, 'html.parser')
      table_contents = bs3.find_all('table')
  
      workbook = openpyxl.Workbook()
      sheet = workbook.active
      sheet['A1'] = "Name"
      sheet['B1'] = "Link"
  
      row = 2  # Start from row 2 to skip the header row
      for table in table_contents:
          list_of_links = table.find_all('a', href=re.compile(r"^/en/players/[a-f0-9]{8}/[A-Za-z-]+$"))
          for link in list_of_links:
              name = link['href'].split('/')[-1].replace('-', ' ')
              href = link['href']
              sheet.cell(row=row, column=1, value=name)
              sheet.cell(row=row, column=2, value=href)
              row += 1
  
      workbook.save('player_profiles.xlsx')
getReports()




def link_generator(player_name):
    count = 0
    df = pd.read_excel('player_profiles.xlsx')
    b_value = None 
    if player_name in df['Name'].values:
        b_value = df[df['Name'] == player_name]['Link'].values[0]
    return b_value

# generates player data and returns key, value data as separate list
def get_players_data(player_name):
    count = 0
    player_stat_keys_raw = []
    player_stat_values = []
    global player_stat_topics
    player_stat_topics = ['Standard Stats', 'Shooting', 'Passing', 'Pass Types',
                          'Goal and Shot Creation','Defense', 'Possession','Miscellaneous Stats', 'Goalkeeping','Advanced Goalkeeping'] 
    player_data_dict = {}
    individual_components = []
    link_to_player_profile = link_generator(player_name)
    html = urlopen("https://fbref.com/"+link_to_player_profile)

    bs = BeautifulSoup(html, 'html.parser')
    global player_image_url
    player_image_url = bs.find('div', {'class':'media-item'}).find('img').attrs['src']

    scout_report_link = bs.find('div',{'class':'section_heading_text'}).find('a').attrs['href']
    
    scout_html = urlopen("https://fbref.com/"+ scout_report_link)
    
    bs_scout_all = BeautifulSoup(scout_html, 'lxml')
    bs_scout = bs_scout_all.find('div', {'id': re.compile(r'div_scout_full_')})
    stat_tables_keys = bs_scout.find("table", {'id': re.compile(r'scout_full_')} ).find_all('tr')
    stat_tables_p90Percentile = bs_scout.find("table", {'id': re.compile(r'scout_full_')}).find_all('td')

    for list_th in stat_tables_keys:
        if list_th.find('th').get_text() != "":
            if list_th.find('th').get_text() != "Statistic":
                player_stat_keys_raw.append(list_th.find('th').get_text())
  

    for list_tb_n in range(0, len(stat_tables_p90Percentile),2):
        
        if list_tb_n+1 < len(stat_tables_p90Percentile):
            list_tb = stat_tables_p90Percentile[list_tb_n]
            new_list = [list_tb.get_text()]
            new_list.append(unicodedata.normalize("NFKD",stat_tables_p90Percentile[list_tb_n+1].get_text(strip=True)))
            player_stat_values.append(new_list)
    for stat in player_stat_values:
        if "" in stat:
            player_stat_values.remove(stat)
    for stat in player_stat_keys_raw:
        if stat in player_stat_topics:
            player_stat_keys_raw.remove(stat)
    return player_stat_keys_raw, player_stat_values

#takes in list of statistics and their values and plots it in a Pizza Chart
def show_picture(params, values, name_of_player):
    font_normal = FontManager('https://raw.githubusercontent.com/google/fonts/main/apache/roboto/'
                              'Roboto%5Bwdth,wght%5D.ttf')
    font_italic = FontManager('https://raw.githubusercontent.com/google/fonts/main/apache/roboto/'
                              'Roboto-Italic%5Bwdth,wght%5D.ttf')
    font_bold = FontManager('https://raw.githubusercontent.com/google/fonts/main/apache/robotoslab/'
                            'RobotoSlab%5Bwght%5D.ttf')
    
    image = Image.open(urlopen(player_image_url))
    
    # Create a circular mask for the image
    mask = Image.new('L', image.size, 0)
    draw = ImageDraw.Draw(mask)
    draw.ellipse((0, 0) + image.size, fill=255)

    # Apply the mask to the image
    masked_img = ImageOps.fit(image, mask.size, centering=(0.5, 0.5))
    masked_img.putalpha(mask)
    
    # color for the slices and text
    slice_colors = ["#bbEE90"] * 5 + ["#FF93ff"] * 5 + ["#FFCCCB"] * 5 + ["#87CEEB"] * 5
    text_colors = ["#000000"] * 20

    # instantiate PyPizza class
    baker = PyPizza(
        params=params,                  # list of parameters
        background_color="#132257",     # background color
        straight_line_color="#000000",  # color for straight lines
        straight_line_lw=1,             # linewidth for straight lines
        last_circle_color="#000000",    # color for last line
        last_circle_lw=1,               # linewidth of last circle
        other_circle_lw=0,              # linewidth for other circles
        inner_circle_size=11            # size of inner circle
    )


    # plot pizza
    fig, ax = baker.make_pizza(
        values,                          # list of values
        figsize=(10, 12),                # adjust the figsize according to your need
        color_blank_space="same",        # use the same color to fill blank space
        slice_colors=slice_colors,       # color for individual slices
        value_colors=text_colors,        # color for the value-text
        value_bck_colors=slice_colors,   # color for the blank spaces
        blank_alpha=0.4,                 # alpha for blank-space colors
        kwargs_slices=dict(
            edgecolor="#000000", zorder=2, linewidth=1
        ),                               # values to be used when plotting slices
        kwargs_params=dict(
            color="#ffffff", fontsize=13,
            fontproperties=font_bold.prop, va="center"
        ),                               # values to be used when adding parameter labels
        kwargs_values=dict(
            color="#ffffff", fontsize=11,
            fontproperties=font_normal.prop, zorder=3,
            bbox=dict(
                edgecolor="#000000", facecolor="cornflowerblue",
                boxstyle="round,pad=0.2", lw=1
            )
        )                                # values to be used when adding parameter-values labels
    )

    # add title
    fig.text(
        0.515, 0.945, name_of_player, size=27,
        ha="center", fontproperties=font_bold.prop, color="#ffffff"
    )

    # add subtitle
    fig.text(
        0.515, 0.925,
        "Percentile Rank vs Top-Five League Players in their Position for LAST 365 DAYS",
        size=13,
        ha="center", fontproperties=font_bold.prop, color="#ffffff"
    )

    # add credits
    CREDIT_2 = "inspired by: @Worville, @FootballSlices, @somazerofc & @Soumyaj15209314"
    CREDIT_3 = "Automated By @josephnk23"

    fig.text(
        0.99, 0.08, f"\n{CREDIT_2}\n{CREDIT_3}", size=15,
        fontproperties=font_italic.prop, color="#ffffff",
        ha="right"
    )

    # add text
    fig.text(
        0.23, 0.9," Standard        Passing        Possession         Defense", size=18,
        fontproperties=font_bold.prop, color="#ffffff"
    )

    # add rectangles
    fig.patches.extend([
        plt.Rectangle(
            (0.205, 0.9), 0.025, 0.0196, fill=True, color="#bbEE90",
            transform=fig.transFigure, figure=fig
        ),
        plt.Rectangle(
            (0.365, 0.9), 0.025, 0.0196, fill=True, color="#FF93ff",
            transform=fig.transFigure, figure=fig
        ),
        plt.Rectangle(
            (0.505, 0.9), 0.025, 0.0196, fill=True, color="#FFCCCB",
            transform=fig.transFigure, figure=fig
        ),
        plt.Rectangle(
            (0.695, 0.9), 0.025, 0.0196, fill=True, color="#87CEEB",
            transform=fig.transFigure, figure=fig
        ),
    ])

    # add image
    ax_image = add_image(
        masked_img, fig, left=0.472, bottom=0.457, width=0.086, height=0.08, zorder= -1
    )     
   
    plt.savefig(name_of_player +".jpg", format='jpg')
    end_time = time.time()
    time_elapsed = end_time - start_time
    print("Elapsed time: {:.2f} seconds".format(time_elapsed))


#takes input from the user and calls methods to get players data , 'Ben Chilwell', 'Hakim Ziyech', 'Conor Gallagher', 'Reece James', 'Kalidou Koulibaly', 'Cesar Azpilicueta', 'Kai Havertz', 'Carney Chukwuemeka', 'Malang Sarr', 'Emerson Palmieri', 'Moises Caiceido', 'Joao Felix', 'Enzo Fernandez','David Datro Fofana', 'Wesley Fofana'
# and plot it onto the Pizza Chart


name_of_player = input("Enter the name")
global start_time
start_time = time.time()
def stats_gobbler():
    data_scraper = get_players_data(name_of_player)
    
    global params

    params_raw = ['Non-Penalty Goals', 'Assists', 'Goals + Assists', 'Yellow Cards', 'Red Cards', 'Passes Attempted',
                          'Pass Completion %', 'Progressive Passes', 'Through Balls','Key Passes', 'Touches','Take-Ons Attempted', 'Successful Take-Ons', 'Miscontrols',
                          'Dispossessed','Tackles','Tackles Won','Shots Blocked', 'Interceptions', 'Clearances']
    values = []

    params = ['Non-Penalty\nGoals', 'Non-Penalty\nxG', 'Goals+Assists', 'Yellow Cards', 'Red Cards', 'Passes\nAttempted',
                          'Pass\nCompletion %', 'Progressive\nPasses', 'Through Balls','Key Passes', 'Touches','Take-Ons\nAttempted', 'Successful\nTake-Ons', 'Miscontrols',
                          'Dispossessed','Tackles','Tackles Won','Shots Blocked', 'Interceptions', 'Clearances']

#Concatenates the Stat and their values into an accessible Dataframe

    df = pd.concat([pd.Series(data_scraper[0]), pd.Series(data_scraper[1])], axis=1)


    for name in params_raw:
        if name in df[0].tolist():
            value = df.loc[df[0]==name,1].iloc[0]
            values.append(int(value.pop()))
    show_picture(params, values, name_of_player)
stats_gobbler()
